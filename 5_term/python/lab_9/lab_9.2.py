import openpyxl

file = "C:\\Users\\Бобр Даша\\Desktop\\university\\3 КУРС\\5 сем\\Язык Python для работы с данными\\lab_9\\salary.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb.active

data = []
for human in range(2, ws.max_row + 1):
    if ws[f"B{human}"].value != None:
        data.append(
            {
                "Фамилия": ws[f"B{human}"].value,
                "Отдел": ws[f"C{human}"].value,
                "Сумма зарплаты": ws[f"F{human}"].value,
            }
        )

res = []
max_zp = [0, None]
min_zp = [9999999999, None]

for human in data:
    if human["Сумма зарплаты"] > max_zp[0]:
        max_zp[0] = human["Сумма зарплаты"]
        max_zp[1] = human["Фамилия"]
    if human["Сумма зарплаты"] < min_zp[0]:
        min_zp[0] = human["Сумма зарплаты"]
        min_zp[1] = human["Фамилия"]

res.append(max_zp)
res.append(min_zp)

count = summ = i = 0
depart = data[0]["Отдел"]

for i in data:
    if i["Отдел"] == depart:
        summ += i["Сумма зарплаты"]
        count += 1
    else:
        res.append([depart, round(summ/count, 2)])
        depart = i["Отдел"]
        count = 1
        summ = i["Сумма зарплаты"]

res.append([depart, data[-1]["Сумма зарплаты"]])      
for i in range(len(res)):
    for j in range(len(res[i])):
        ws.cell(i+14, j+2, res[i][j])



wb.save(file)
